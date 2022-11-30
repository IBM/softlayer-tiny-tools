#!/usr/bin/python3
# -*- coding: UTF-8 -*-
import SoftLayer
import getopt
import sys
from openpyxl import Workbook
import time 
import json

start_date = ''
end_date = ''
output_format  = ''
invoice_type = []
supported_invoice_type = ("ALL", "NEW", "RECURRING", "ONE-TIME-CHARGE","CREDIT","REFUND","MANUAL_PAYMENT_CREDIT" )

try:
    opts, args = getopt.getopt(
        sys.argv[1:],
        "hs:e:f:t:",
        ["start_date=", "end_date=", "output_format=", "type="],
        )
except getopt.GetoptError:
    print("%s -s <start_date> -e <end_date> -f <pdf_or_xls> -t <invoice-type>" % sys.argv[0])
    print("example:")
    print("         %s -s 03/01/2022 -e 03/30/2022 -f pdf" % sys.argv[0])
    print("")
    sys.exit(2)
for opt, arg in opts:
    if opt == '-h':
        print("%s -s <start_date> -e <end_date> -f <pdf_or_xls> -t <invoice-type>" % sys.argv[0])
        print("example:")
        print("         %s -s 03/01/2022 -e 03/30/2022 -f pdf -t ALL" % sys.argv[0])
        print("")
        sys.exit()
    elif opt in ("-s", "--start_date"):
        start_date = arg
    elif opt in ("-e", "--end_date"):
        end_date = arg
    elif opt in ("-f", "--output_format"):
        output_format = arg
    elif opt in ("-t", "--type"):
        if arg not in supported_invoice_type:
            print("unsupported invoice type")
            exit(1)
        invoice_type.append(arg)


class SL_Service():
    def __init__(self):
        self.client = SoftLayer.create_client_from_env()
        self.account_service = self.client['Account']
        self.billing_invoice_service = self.client['Billing_Invoice']
    
    # get all of invoices during a times 
    # format of start and end date is  '01/01/2015'
    def get_invoices(self, startDate, endDate):
        objectMask = "mask[id, createDate, typeCode, amount, invoiceTotalAmount, invoiceTotalOneTimeAmount, invoiceTotalOneTimeTaxAmount, invoiceTotalPreTaxAmount, invoiceTotalRecurringAmount, invoiceTotalRecurringTaxAmount, payment, startingBalance, endingBalance, items[associatedInvoiceItemId, billingItemId, categoryCode, createDate, description, domainName, hostName, hourlyRecurringFee, id, invoiceId, laborAfterTaxAmount, laborFee, laborFeeTaxRate, laborTaxAmount, notes, oneTimeAfterTaxAmount, oneTimeFee, oneTimeFeeTaxRate, oneTimeTaxAmount, parentId, productItemId, recurringAfterTaxAmount, recurringFee, recurringFeeTaxRate, recurringTaxAmount, resourceTableId, serviceProviderId, setupAfterTaxAmount, setupFee, setupFeeDeferralMonths, setupFeeTaxRate, setupTaxAmount]]"
        objectFilter = {
            'invoices': {
                'createDate': {
                    'operation': 'betweenDate',
                    'options': [
                        {
                        'name': 'startDate',
                        'value': [startDate]
                        },
                        {
                        'name': 'endDate',
                        'value': [endDate]
                        }
                    ]
                }
            }
        }

        if "ALL" not in invoice_type:
            objectFilter["invoices"]["typeCode"] =  {
                        'operation': 'in', 
                        'options': [{
                            'name': 'data',
                            'value': []
                        }]
                }
            for it in invoice_type:
                objectFilter["invoices"]['typeCode']['options'][0]['value'].append(it)

        try:
            invoices = self.account_service.getInvoices(mask=objectMask, filter=objectFilter)
           # print(invoices)
            return  invoices
        except SoftLayer.SoftLayerAPIError as e:
            raise e
           # raise ("Failed to get invoice list with error code %s and err msg %s " % (e.faultCode, e.faultString))

    def save_invoices(self, invoices, format):
        wb = Workbook()
        dest_filename = 'summary_{}.xlsx'.format(time.strftime('%Y年%m月%d日%H时%M分%S秒'))

        ws1 = wb.active
        ws1.title = "range names"

        table_header = [
            "id", 
            "createDate", 
            "typeCode",
            "amount", 
            "invoiceTotalAmount",
            "invoiceTotalOneTimeAmount",
            "invoiceTotalOneTimeTaxAmount",
            "invoiceTotalPreTaxAmount",
            "invoiceTotalRecurringAmount",
            "invoiceTotalRecurringTaxAmount",
            "payment",
            "startingBalance",
            "endingBalance"
            ]
        table_header_output = [
            "账单编号", 
            "创建日期", 
            "账单类型",
            "总价", 
            "发票总金额",
            "invoiceTotalOneTimeAmount",
            "invoiceTotalOneTimeTaxAmount",
            "invoiceTotalPreTaxAmount",
            "invoiceTotalRecurringAmount",
            "invoiceTotalRecurringTaxAmount",
            "payment",
            "期初",
            "期末"
            ]

        ws1.append(table_header_output)

        row = 2 
        for invoice in invoices:
            col = 1
            for key in table_header:
                if key not in invoice.keys():
                     ws1.cell(column=col, row=row, value="不可用")
                else:
                    ws1.cell(column=col, row=row, value=str(invoice[key]))
                col = col+1
            row = row +1 
        
        wb.save(filename = dest_filename)

if __name__ == "__main__":
    service = SL_Service()
    invoices = service.get_invoices(start_date, end_date)
    print(json.dumps(invoices, sort_keys=True, indent=2, separators=(',', ': ')))
    if len(invoices) != 0: 
        service.save_invoices(invoices, output_format)
    else: 
        print("没有获取到任何账单，请调整日期范围并检查api key 权限是否设置正确。")